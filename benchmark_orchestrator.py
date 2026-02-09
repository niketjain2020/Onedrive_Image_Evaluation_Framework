#!/usr/bin/env python3
"""
Restyle Benchmark Orchestrator v1.1.0

This module provides validation and utility functions for the benchmark pipeline.
The main execution is driven by Claude Code using Playwright MCP tools.

Usage:
    python benchmark_orchestrator.py validate        # Run Phase 1 validation
    python benchmark_orchestrator.py init            # Initialize output directory
    python benchmark_orchestrator.py synthesize      # Run Phase 6 synthesis
    python benchmark_orchestrator.py persist         # Run Phase 7 persist to Excel
    python benchmark_orchestrator.py compare         # Compare current run to baseline
    python benchmark_orchestrator.py build-prompt <style>
    python benchmark_orchestrator.py compute-rankings
"""

import json
import os
import shutil
import sys
import uuid
from datetime import datetime
from pathlib import Path
from typing import Any

from config import (
    PROJECT_ROOT, RUNS_DIR, EXCEL_PATH,
    STYLE_ASSERTIONS_PATH, ACRUE_V3_PROMPT_PATH, RUN_SPEC_PATH,
    GEMINI_MODEL,
)

# Excel support
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel features disabled.")

# Comparison support
try:
    from compare_runs import compare_runs, print_comparison_console, save_comparison, find_latest_prior_run
    COMPARE_AVAILABLE = True
except ImportError:
    COMPARE_AVAILABLE = False

# Base paths (kept for backward compat; prefer config.py imports)
SCRIPT_DIR = PROJECT_ROOT


class ValidationError(Exception):
    """Raised when validation fails."""
    pass


def load_run_spec() -> dict:
    """Load and return the run specification."""
    if not RUN_SPEC_PATH.exists():
        raise ValidationError(f"V1 FAIL: run_spec.json not found at {RUN_SPEC_PATH}")

    try:
        with open(RUN_SPEC_PATH, 'r') as f:
            return json.load(f)
    except json.JSONDecodeError as e:
        raise ValidationError(f"V2 FAIL: run_spec.json is not valid JSON: {e}")


def validate_phase1() -> dict:
    """
    Phase 1: VALIDATE

    All preconditions must pass before proceeding.
    Returns the validated run_spec on success.
    """
    errors = []

    # V1: run_spec.json exists
    if not RUN_SPEC_PATH.exists():
        errors.append("V1 FAIL: run_spec.json does not exist")
        # Can't continue without spec
        raise ValidationError("\n".join(errors))

    # V2: Valid JSON
    try:
        spec = load_run_spec()
    except ValidationError as e:
        errors.append(str(e))
        raise ValidationError("\n".join(errors))

    # V3: styles array has exactly 3 items
    styles = spec.get("styles", [])
    if len(styles) != 3:
        errors.append(f"V3 FAIL: styles must have exactly 3 items, got {len(styles)}")

    # V4: image_count equals 3
    image_count = spec.get("image_count", 0)
    if image_count != 3:
        errors.append(f"V4 FAIL: image_count must be 3, got {image_count}")

    # V5: output_dir does NOT exist
    output_dir = SCRIPT_DIR / spec.get("output_dir", "")
    if output_dir.exists():
        errors.append(f"V5 FAIL: output_dir already exists: {output_dir}")

    # V6: If baseline_run_id is set, that directory must exist
    baseline_run_id = spec.get("baseline_run_id")
    if baseline_run_id:
        baseline_dir = RUNS_DIR / baseline_run_id
        if not baseline_dir.exists():
            errors.append(f"V6 FAIL: baseline_run_id directory not found: {baseline_dir}")

    # V7: GEMINI_API_KEY is set
    if not os.environ.get("GEMINI_API_KEY"):
        errors.append("V7 FAIL: GEMINI_API_KEY environment variable not set")

    # V8: Playwright MCP check - this is validated externally by Claude Code
    # We skip this in Python validation

    if errors:
        raise ValidationError("\n".join(errors))

    print("Phase 1 VALIDATE: All checks passed")
    print(f"  Run ID: {spec['run_id']}")
    print(f"  Styles: {', '.join(styles)}")
    print(f"  Images: {image_count}")
    print(f"  Output: {output_dir}")

    return spec


def init_output_dir(spec: dict) -> Path:
    """
    Initialize the output directory structure.

    Creates:
        {output_dir}/
        {output_dir}/originals/
        {output_dir}/restyled/
    """
    output_dir = SCRIPT_DIR / spec["output_dir"]

    # Create directory structure
    (output_dir / "originals").mkdir(parents=True, exist_ok=True)
    (output_dir / "restyled").mkdir(parents=True, exist_ok=True)

    # Copy run_spec.json
    with open(output_dir / "run_spec.json", 'w') as f:
        json.dump(spec, f, indent=2)

    print(f"Initialized output directory: {output_dir}")
    return output_dir


def get_style_slug(style: str) -> str:
    """Convert style name to slug format."""
    return style.lower().replace(" ", "_")


def load_style_assertions(style: str) -> dict:
    """Load assertions for a specific style."""
    with open(STYLE_ASSERTIONS_PATH, 'r') as f:
        data = json.load(f)

    styles = data.get("styles", {})
    if style in styles:
        return styles[style]

    # Return generic assertions if style not found
    print(f"Warning: No specific assertions for '{style}', using generic")
    return {
        "description": f"{style} artistic style",
        "assertions": {
            "accuracy": [
                f"Does the output exhibit characteristics of {style} style?",
                "Is the subject recognizable from the original?",
                "Does the style transformation feel authentic?",
                "Are key visual elements preserved?",
                "Does the quality match professional work?"
            ],
            "completeness": [
                "Is the style applied to the entire image?",
                "Is the background appropriately styled?",
                "Are all subjects structurally intact?",
                "Is style saturation appropriate?",
                "Are all elements stylistically consistent?"
            ],
            "relevance": [
                f"Does the output match {style} aesthetic?",
                "Is the style application appropriate?",
                "Does the mood match the intended style?",
                "Are style elements authentic?"
            ],
            "usefulness": [
                "Is the image suitable for sharing?",
                "Is it free of obvious artifacts?",
                "Is the subject clearly visible?",
                "Is resolution sufficient?"
            ],
            "exceptional": [
                "Does it look professionally made?",
                "Does it evoke positive emotions?",
                "Would users want to share it?",
                "Does it have 'wow factor'?",
                "Does it stand out from typical AI outputs?"
            ]
        }
    }


def build_acrue_prompt(style: str, style_data: dict) -> str:
    """Build the ACRUE v3 evaluation prompt for a specific style."""
    with open(ACRUE_V3_PROMPT_PATH, 'r') as f:
        template = f.read()

    assertions = style_data.get("assertions", {})

    def format_assertions(dimension: str) -> str:
        items = assertions.get(dimension, [])
        lines = []
        for i, q in enumerate(items, 1):
            prefix = dimension[0].upper()
            lines.append(f"{prefix}{i}. {q}")
        return "\n".join(lines)

    prompt = template.replace("{STYLE_NAME}", style)
    prompt = prompt.replace("{STYLE_DESCRIPTION}", style_data.get("description", ""))
    prompt = prompt.replace("{ASSERTIONS_ACCURACY}", format_assertions("accuracy"))
    prompt = prompt.replace("{ASSERTIONS_COMPLETENESS}", format_assertions("completeness"))
    prompt = prompt.replace("{ASSERTIONS_RELEVANCE}", format_assertions("relevance"))
    prompt = prompt.replace("{ASSERTIONS_USEFULNESS}", format_assertions("usefulness"))
    prompt = prompt.replace("{ASSERTIONS_EXCEPTIONAL}", format_assertions("exceptional"))

    return prompt


def compute_feasibility_rankings(acrue_results: list) -> dict:
    """
    Compute feasibility rankings from ACRUE v3 evaluation results.

    Groups scores by style and ranks by average weighted total.
    """
    # Group by style
    style_scores = {}
    for result in acrue_results:
        style = result["style"]
        if style not in style_scores:
            style_scores[style] = []

        # Get the summary scores
        summary = result.get("summary", {})
        weighted_total = summary.get("weighted_total", result.get("total", 0))
        percentage = summary.get("percentage", result.get("percentage", 0))
        grade = summary.get("grade", result.get("grade", "F"))

        style_scores[style].append({
            "weighted_total": weighted_total,
            "percentage": percentage,
            "grade": grade,
            "dimensions": result.get("dimensions", {})
        })

    # Calculate averages and build rankings
    rankings_data = []
    for style, scores in style_scores.items():
        avg_score = sum(s["weighted_total"] for s in scores) / len(scores)
        avg_pct = sum(s["percentage"] for s in scores) / len(scores)

        # Determine average grade
        grade_values = {"A+": 5, "A": 4, "B": 3, "C": 2, "F": 1}
        avg_grade_val = sum(grade_values.get(s["grade"], 1) for s in scores) / len(scores)
        if avg_grade_val >= 4.5:
            avg_grade = "A+"
        elif avg_grade_val >= 3.5:
            avg_grade = "A"
        elif avg_grade_val >= 2.5:
            avg_grade = "B"
        elif avg_grade_val >= 1.5:
            avg_grade = "C"
        else:
            avg_grade = "F"

        # Build reasoning from dimension breakdowns
        dim_summaries = []
        for dim in ["accuracy", "completeness", "relevance", "usefulness", "exceptional"]:
            dim_scores = [s["dimensions"].get(dim, {}).get("weighted_score", 0)
                         for s in scores if "dimensions" in s]
            if dim_scores:
                dim_avg = sum(dim_scores) / len(dim_scores)
                dim_summaries.append(f"{dim.title()}: {dim_avg:.1f}")

        reasoning = f"Average across {len(scores)} images. " + ", ".join(dim_summaries)

        rankings_data.append({
            "style": style,
            "avg_score": round(avg_score, 2),
            "avg_percentage": round(avg_pct, 1),
            "avg_grade": avg_grade,
            "reasoning": reasoning
        })

    # Sort by avg_score descending (highest = rank 1)
    rankings_data.sort(key=lambda x: x["avg_score"], reverse=True)

    # Assign ranks
    rankings = []
    for i, data in enumerate(rankings_data, 1):
        rankings.append({
            "rank": i,
            **data
        })

    return {
        "judge": f"{GEMINI_MODEL} (ACRUE v3)",
        "timestamp": datetime.now().isoformat(),
        "methodology": "Rankings computed from ACRUE v3 weighted_total averages",
        "rankings": rankings
    }


def synthesize_results(spec: dict, gemini_rankings: dict, opus_rankings: dict) -> dict:
    """
    Phase 6: SYNTHESIS

    Combines feasibility and preference rankings using weighted formula.
    """
    feasibility_weight = spec["synthesis"]["feasibility_weight"]
    preference_weight = spec["synthesis"]["preference_weight"]

    # Build lookup for ranks
    gemini_ranks = {r["style"]: r["rank"] for r in gemini_rankings["rankings"]}
    opus_ranks = {r["style"]: r["rank"] for r in opus_rankings["rankings"]}

    # Calculate final scores
    final_scores = []
    for style in spec["styles"]:
        g_rank = gemini_ranks.get(style, 3)
        o_rank = opus_ranks.get(style, 3)
        final_score = (feasibility_weight * g_rank) + (preference_weight * o_rank)
        final_scores.append({
            "style": style,
            "gemini_rank": g_rank,
            "opus_rank": o_rank,
            "final_score": final_score
        })

    # Sort by final_score (lowest = winner)
    final_scores.sort(key=lambda x: x["final_score"])

    # Assign ranks
    for i, item in enumerate(final_scores, 1):
        item["rank"] = i

    winner = final_scores[0]["style"]

    return {
        "winner": winner,
        "rankings": final_scores,
        "weights": {
            "feasibility": feasibility_weight,
            "preference": preference_weight
        }
    }


def generate_report(spec: dict, gemini: dict, opus: dict, synthesis: dict,
                    comparison: dict = None) -> str:
    """Generate the final report.md content, optionally including comparison."""
    run_id = spec["run_id"]

    lines = [
        f"# Benchmark Run: {run_id}",
        "",
        "## Final Rankings",
        "",
        "| Rank | Style | Gemini Rank | Opus Rank | Final Score |",
        "|------|-------|-------------|-----------|-------------|"
    ]

    for r in synthesis["rankings"]:
        lines.append(f"| {r['rank']} | {r['style']} | {r['gemini_rank']} | {r['opus_rank']} | {r['final_score']:.2f} |")

    lines.extend([
        "",
        f"**Winner: {synthesis['winner']}**",
        "",
        "## Gemini Assessment (Feasibility)",
        ""
    ])

    for r in gemini["rankings"]:
        lines.append(f"### {r['rank']}. {r['style']} (Score: {r['avg_score']}/25.0, {r['avg_percentage']:.1f}%, Grade: {r['avg_grade']})")
        lines.append(f"{r['reasoning']}")
        lines.append("")

    lines.extend([
        "## Opus Assessment (Preference)",
        ""
    ])

    for r in opus["rankings"]:
        lines.append(f"### {r['rank']}. {r['style']} (Appeal: {r.get('appeal_score', 'N/A')})")
        lines.append(f"{r['reasoning']}")
        lines.append("")

    # --- Comparison section (new in v1.1) ---
    if comparison and "styles" in comparison:
        lines.extend([
            "## Comparison vs Baseline",
            "",
            f"Baseline: **{comparison.get('baseline_run_id', 'N/A')}**",
            "",
            "| Style | Baseline | Current | Delta | Status |",
            "|-------|----------|---------|-------|--------|",
        ])
        for s in comparison["styles"]:
            sign = "+" if s["delta_pct"] > 0 else ""
            lines.append(
                f"| {s['style']} | {s['baseline_pct']:.1f}% | {s['current_pct']:.1f}% "
                f"| {sign}{s['delta_pct']:.1f}% | {s['status']} |"
            )
        summary_c = comparison.get("summary", {})
        lines.extend([
            "",
            f"**Verdict: {summary_c.get('overall_verdict', 'N/A')}** "
            f"({summary_c.get('improved', 0)} improved, "
            f"{summary_c.get('regressed', 0)} regressed, "
            f"{summary_c.get('unchanged', 0)} unchanged)",
            "",
        ])

    lines.extend([
        "## Metadata",
        "",
        f"- **Run ID**: {run_id}",
        f"- **Pipeline Version**: {spec['pipeline_version']}",
        f"- **ACRUE Version**: {spec['acrue_version']}",
        f"- **Gemini Model**: {spec['judges']['feasibility']['model']}",
        f"- **Opus Model**: {spec['judges']['preference']['model']}",
        f"- **Timestamp**: {datetime.now().isoformat()}",
        f"- **Styles Tested**: {', '.join(spec['styles'])}",
        f"- **Images Per Style**: {spec['image_count']}",
        f"- **Total Evaluations**: {spec['image_count'] * len(spec['styles'])}",
        ""
    ])

    return "\n".join(lines)


def _is_valid_xlsx(path: Path) -> bool:
    """Check whether a file is a valid OOXML .xlsx (ZIP-based), not OLE2 .xls."""
    if not path.exists():
        return False
    try:
        with open(path, "rb") as f:
            magic = f.read(4)
        # OLE2 magic: D0 CF 11 E0   (old .xls format)
        if magic == b"\xd0\xcf\x11\xe0":
            return False
        # Try loading with openpyxl
        load_workbook(path)
        return True
    except Exception:
        return False


def _create_summary_headers(ws):
    """Write header row on a Summary worksheet."""
    headers = [
        "run_id", "timestamp", "pipeline_version", "acrue_version",
        "styles", "image_count", "winner", "gemini_top", "opus_top",
        "artifacts_path", "vs_baseline", "regressions", "improvements",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")


def persist_to_excel(spec: dict, synthesis: dict, comparison: dict = None):
    """
    Phase 7: PERSIST

    Append results to ai_restyle_benchmark.xlsx.
    Handles corrupted OLE2 files by backing up and recreating.
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not available, skipping Excel persist")
        return

    run_id = spec["run_id"]

    # --- Handle corrupted/wrong-format file ---
    if EXCEL_PATH.exists() and not _is_valid_xlsx(EXCEL_PATH):
        backup = EXCEL_PATH.with_suffix(".xls.bak")
        print(f"Warning: {EXCEL_PATH.name} is not a valid .xlsx file (likely OLE2 .xls).")
        print(f"  Backing up to {backup.name} and creating a fresh .xlsx")
        shutil.copy2(EXCEL_PATH, backup)
        EXCEL_PATH.unlink()

    # Load or create workbook
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Summary"
        _create_summary_headers(ws)

    # Get or create summary sheet
    if "Summary" in wb.sheetnames:
        summary_ws = wb["Summary"]
    else:
        summary_ws = wb.create_sheet("Summary", 0)
        _create_summary_headers(summary_ws)

    # Ensure header row has the new comparison columns (migration)
    if summary_ws.cell(row=1, column=11).value is None:
        summary_ws.cell(row=1, column=11, value="vs_baseline").font = Font(bold=True)
        summary_ws.cell(row=1, column=12, value="regressions").font = Font(bold=True)
        summary_ws.cell(row=1, column=13, value="improvements").font = Font(bold=True)

    # Add summary row
    next_row = summary_ws.max_row + 1
    gemini_top = None
    opus_top = None

    # Find actual tops
    for r in synthesis["rankings"]:
        if r["gemini_rank"] == 1:
            gemini_top = r["style"]
        if r["opus_rank"] == 1:
            opus_top = r["style"]

    # Comparison columns
    vs_baseline = ""
    regressions = 0
    improvements = 0
    if comparison and "summary" in comparison:
        vs_baseline = comparison.get("baseline_run_id", "")
        regressions = comparison["summary"].get("regressed", 0)
        improvements = comparison["summary"].get("improved", 0)

    summary_data = [
        run_id,
        datetime.now().isoformat(),
        spec["pipeline_version"],
        spec["acrue_version"],
        ", ".join(spec["styles"]),
        spec["image_count"],
        synthesis["winner"],
        gemini_top,
        opus_top,
        spec["output_dir"],
        vs_baseline,
        regressions,
        improvements,
    ]

    for col, value in enumerate(summary_data, 1):
        summary_ws.cell(row=next_row, column=col, value=value)

    # Create detailed sheet for this run
    if run_id in wb.sheetnames:
        del wb[run_id]

    run_ws = wb.create_sheet(run_id)

    # Header
    run_ws.cell(row=1, column=1, value=f"Benchmark Run: {run_id}").font = Font(bold=True, size=14)
    run_ws.cell(row=2, column=1, value=f"Winner: {synthesis['winner']}").font = Font(bold=True)

    # Rankings table
    run_ws.cell(row=4, column=1, value="Final Rankings").font = Font(bold=True)
    headers = ["Rank", "Style", "Gemini Rank", "Opus Rank", "Final Score"]
    for col, header in enumerate(headers, 1):
        cell = run_ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")

    for i, r in enumerate(synthesis["rankings"], 6):
        run_ws.cell(row=i, column=1, value=r["rank"])
        run_ws.cell(row=i, column=2, value=r["style"])
        run_ws.cell(row=i, column=3, value=r["gemini_rank"])
        run_ws.cell(row=i, column=4, value=r["opus_rank"])
        run_ws.cell(row=i, column=5, value=r["final_score"])

    # Save
    wb.save(EXCEL_PATH)
    print(f"Persisted to Excel: {EXCEL_PATH}")


def run_comparison(spec: dict, output_dir: Path) -> dict | None:
    """Auto-compare the current run against the most recent prior run."""
    if not COMPARE_AVAILABLE:
        print("Warning: compare_runs module not available, skipping comparison")
        return None

    run_id = spec["run_id"]

    # Use explicit baseline from run_spec, or find latest prior run
    baseline_id = spec.get("baseline_run_id")
    if not baseline_id:
        baseline_id = find_latest_prior_run(run_id)

    if not baseline_id:
        print("No baseline run found for comparison (first run?)")
        return None

    print(f"\nComparing {run_id} vs baseline {baseline_id}...")
    try:
        comparison = compare_runs(run_id, baseline_id)
        print_comparison_console(comparison)
        save_comparison(comparison, output_dir)
        return comparison
    except Exception as e:
        print(f"Warning: Comparison failed: {e}")
        return None


def main():
    """Main entry point."""
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    command = sys.argv[1].lower()

    try:
        if command == "validate":
            spec = validate_phase1()
            print(json.dumps(spec, indent=2))

        elif command == "init":
            spec = load_run_spec()
            output_dir = init_output_dir(spec)
            print(f"OUTPUT_DIR={output_dir}")

        elif command == "synthesize":
            spec = load_run_spec()
            output_dir = SCRIPT_DIR / spec["output_dir"]

            # Load judge results
            with open(output_dir / "gemini.json", 'r') as f:
                gemini = json.load(f)
            with open(output_dir / "opus.json", 'r') as f:
                opus = json.load(f)

            # Synthesize
            synthesis = synthesize_results(spec, gemini, opus)

            # Auto-compare to baseline
            comparison = run_comparison(spec, output_dir)

            # Generate report (with comparison if available)
            report = generate_report(spec, gemini, opus, synthesis, comparison)

            # Save
            with open(output_dir / "synthesis.json", 'w') as f:
                json.dump(synthesis, f, indent=2)
            with open(output_dir / "report.md", 'w') as f:
                f.write(report)

            print(f"Winner: {synthesis['winner']}")
            print(f"Report saved to: {output_dir / 'report.md'}")

        elif command == "persist":
            spec = load_run_spec()
            output_dir = SCRIPT_DIR / spec["output_dir"]

            with open(output_dir / "synthesis.json", 'r') as f:
                synthesis = json.load(f)

            # Load comparison if it exists
            comparison = None
            comp_path = output_dir / "comparison.json"
            if comp_path.exists():
                with open(comp_path, 'r') as f:
                    comparison = json.load(f)

            persist_to_excel(spec, synthesis, comparison)

        elif command == "compare":
            spec = load_run_spec()
            output_dir = SCRIPT_DIR / spec["output_dir"]
            comparison = run_comparison(spec, output_dir)
            if comparison is None:
                print("No comparison generated.")
                sys.exit(1)

        elif command == "build-prompt":
            # Build ACRUE prompt for a style
            if len(sys.argv) < 3:
                print("Usage: benchmark_orchestrator.py build-prompt <style>")
                sys.exit(1)

            style = sys.argv[2]
            style_data = load_style_assertions(style)
            prompt = build_acrue_prompt(style, style_data)
            print(prompt)

        elif command == "compute-rankings":
            # Compute rankings from acrue.json
            spec = load_run_spec()
            output_dir = SCRIPT_DIR / spec["output_dir"]

            with open(output_dir / "acrue.json", 'r') as f:
                acrue_results = json.load(f)

            rankings = compute_feasibility_rankings(acrue_results)

            with open(output_dir / "gemini.json", 'w') as f:
                json.dump(rankings, f, indent=2)

            print(json.dumps(rankings, indent=2))

        else:
            print(f"Unknown command: {command}")
            print(__doc__)
            sys.exit(1)

    except ValidationError as e:
        print(f"VALIDATION FAILED:\n{e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
