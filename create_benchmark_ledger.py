"""
Create the AI Restyle Benchmark Excel ledger.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_benchmark_ledger():
    """Create the initial benchmark ledger Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Define headers
    headers = [
        "run_id",
        "timestamp",
        "pipeline_version",
        "acrue_version",
        "styles",
        "image_count",
        "winner",
        "gemini_top",
        "opus_top",
        "artifacts_path"
    ]

    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Set column widths
    column_widths = {
        "A": 30,  # run_id
        "B": 22,  # timestamp
        "C": 16,  # pipeline_version
        "D": 14,  # acrue_version
        "E": 35,  # styles
        "F": 12,  # image_count
        "G": 15,  # winner
        "H": 15,  # gemini_top
        "I": 15,  # opus_top
        "J": 45,  # artifacts_path
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save the workbook
    output_path = os.path.join(os.path.dirname(__file__), "ai_restyle_benchmark.xlsx")
    wb.save(output_path)
    print(f"Created benchmark ledger at: {output_path}")
    return output_path

if __name__ == "__main__":
    create_benchmark_ledger()
